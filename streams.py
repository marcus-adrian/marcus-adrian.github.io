import yaml
import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
import time
import sys
from tqdm import tqdm
YELLOW_HIGHLIGHT = PatternFill(start_color='ffff00',
				   end_color='ffff00',
				   fill_type='solid')
GREEN_HIGHLIGHT = PatternFill(start_color='90ee90',
				   end_color='90ee90',
				   fill_type='solid')
ORANGE_HIGHLIGHT = PatternFill(start_color='ffa500',
				   end_color='ffa500',
				   fill_type='solid')

#Retrieves data from the YAML file and returns a dictionary with the vals.
def get_config_variables():
	with open('config.yaml') as f:
		data = yaml.load(f, Loader=yaml.FullLoader)
		return data

def copy_worksheet(workbook, sheetName):
	source = workbook.active
	target = workbook.copy_worksheet(source)
	target.title = sheetName
	return target

def find_row_with_key(worksheet, key):
	rowNum = 0 #loop through and find row for key
	for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=1, values_only=True):
		rowNum += 1
		for cell in row:
			if(cell != None and isinstance(cell, str)):
				if(key in cell):
				   return rowNum
	return 0

def removeRowsBelow(worksheet):
	rowNum = 0 #loop through and find row for mass flows
	massFlowArr = []
	for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=1, values_only=True):
		rowNum += 1
		for cell in row:
			if(cell == "Mass Flows"):
			   massFlowArr.append(rowNum) #More than one intance of mass flow in sheet
	worksheet.delete_rows(massFlowArr[0] + 1, worksheet.max_row) #Delete from first mass flow cell down

def removeZeroRows(worksheet):
	maxRow = worksheet.max_row
	maxCol = worksheet.max_column
	rowNum = 3
	
	for row in worksheet.iter_rows(min_col=3, max_col=maxCol, min_row=3, max_row=maxRow, values_only=True):
		allString = False
		zeroFlag = False #Assume all zero
		for item in row:
			if(isinstance(item, (float, int))): #Is a number?
				if(item != None):
					if(item != 0):
						zeroFlag = True
		for item in row:
			if (isinstance(item, str)):
				allString = True
		if(zeroFlag == False and allString == False):
			worksheet.delete_rows(rowNum, 1)
		rowNum += 1

def addTitle(worksheet, title):
	worksheet.insert_rows(1)
	worksheet['A1'] = title

def addInOutRows(worksheet):
	worksheet.insert_rows(2)
	worksheet['A2'] = "Section In"
	worksheet.insert_rows(3)
	worksheet['A3'] = "Section Out"
	worksheet.merge_cells('A2:CO2') #WHAT A CRAZY FIX WHY
	worksheet.unmerge_cells('A2:CO2') #Three hours of my life gone

def entropyCalculations(worksheet):
	rowIdx = 0
	maxRow = worksheet.max_row
	maxCol = worksheet.max_column
	flag = False
	for row in worksheet.iter_rows(min_row=1, min_col=1,
		max_col=1, max_row=maxRow, values_only=True):
		rowIdx += 1
		for item in row:
			if(item == "Enthalpy Flow"):
				flag = True #found the one row I needed
				newRow = rowIdx + 1
				worksheet.insert_rows(newRow, 2) #Create new row below Enthalpy
				titleCell = "A" + str(newRow)
				unitCell = "B" + str(newRow)
				worksheet[titleCell] = "Entropy Flow" #Add new name for row
				worksheet[unitCell] = "kW/K"
				titleCell = "A" + str(newRow + 1)
				unitCell = "B" + str(newRow + 1)
				worksheet[titleCell] = "Exergy Flow" #Add new name for row
				worksheet[unitCell] = "MW"
		if(flag):
			break

	conversionFactor = 3600 * 1000

	stream_molar_flow = find_row_with_key(worksheet, "Mole Flows")
	stream_molar_flow_units = "B" + str(stream_molar_flow)
	stream_molar_entropy = find_row_with_key(worksheet, "Molar Entropy")
	stream_molar_entropy_units = "B" + str(stream_molar_entropy)

	for row_cells in worksheet.iter_rows(min_row=newRow, max_row=newRow,
		min_col=3, max_col=maxCol):
		for cell in row_cells:
			firstValue = worksheet[str(cell.column_letter) + str(stream_molar_entropy)]
			secondValue = worksheet[str(cell.column_letter) + str(stream_molar_flow)]
			if(firstValue.value != None and secondValue.value != None):
				if(worksheet[stream_molar_flow_units].value == "kmol/hr" and 
					worksheet[stream_molar_entropy_units].value == "J/kmol-K"):
					calculatedValue = (firstValue.value * secondValue.value)/conversionFactor
				else:
					calculatedValue = 1
					print("Unit error in calculating Entropy Flow, required: \n")
					print("Mole Flow Rate: kmol/hr")
					print("Entropy Mixture: J/kmol-K")
				cell.value = calculatedValue
			#Exergy Calculations
			firstValue = worksheet[str(cell.column_letter) + str(find_row_with_key(worksheet, "Enthalpy Flow"))]
			secondValue = worksheet[str(cell.column_letter) + str(find_row_with_key(worksheet, "Entropy Flow"))]
			if(firstValue.value != None and secondValue.value != None):
				calculatedValue = firstValue.value - (0.3 * secondValue.value)
				cell.offset(row=1).value = calculatedValue

	worksheet.delete_rows(find_row_with_key(worksheet, "Description"), 1)

#Returns column letter of first blank after to and from rows
def find_blank(worksheet):
	to_idx = find_row_with_key(worksheet, "To")
	from_idx = find_row_with_key(worksheet, "From")

	for col in worksheet.iter_cols(min_row=to_idx, max_row=to_idx, min_col=3, max_col=worksheet.max_column):
		for cell in col:
			to_cell = str(cell.column_letter) + str(to_idx)
			from_cell = str(cell.column_letter) + str(from_idx)
			if worksheet[to_cell].value == None and worksheet[from_cell].value == None:
				return cell.column_letter
				#worksheet.delete_cols(cell.column_letter,1)

def removeColumns(worksheet):
	to_idx = find_row_with_key(worksheet, "To")
	from_idx = find_row_with_key(worksheet, "From")
	endCol = column_index_from_string(find_blank(worksheet))
	delArray = []
	for col in worksheet.iter_cols(min_row=to_idx, max_row=to_idx, min_col=3, max_col=endCol):
		for cell in col:
			to_cell = str(cell.column_letter) + str(to_idx)
			from_cell = str(cell.column_letter) + str(from_idx)
			if worksheet[to_cell].value != None and worksheet[from_cell].value != None:
				delArray.append(column_index_from_string(cell.column_letter))
	for i in reversed(delArray):
		worksheet.delete_cols(i,1)
	
def addInOutValues(worksheet):
	to_idx = find_row_with_key(worksheet, "To")
	from_idx = find_row_with_key(worksheet, "From")
	endCol = column_index_from_string(find_blank(worksheet))
	for col in worksheet.iter_cols(min_row=to_idx, max_row=to_idx, min_col=3, max_col=endCol):
		for cell in col:
			write_cell = str(cell.column_letter) + str(1)
			to_cell = str(cell.column_letter) + str(to_idx)
			from_cell = str(cell.column_letter) + str(from_idx)
			if worksheet[to_cell].value != None:
				worksheet[write_cell] = "In"
			if worksheet[from_cell].value != None:
				worksheet[write_cell] = "Out"

	worksheet.delete_cols(endCol, worksheet.max_column) 

def calculate_balance(worksheet):
	enthalpy_flow = find_row_with_key(worksheet, "Enthalpy Flow")
	enthalpy_sum = 0 #Initialize counter
	for col in worksheet.iter_cols(min_row=enthalpy_flow, max_row=enthalpy_flow, min_col=3, max_col=worksheet.max_column):
		lastCol = col
		for cell in col:
			in_out_row = str(cell.column_letter) + "1"
			if worksheet[in_out_row].value == "In": #Check if the first row in this col has In or Out
				enthalpy_sum += cell.value 			#If In, Add to the sum
			if worksheet[in_out_row].value == "Out":
				enthalpy_sum -= cell.value 			#If out, Subtract from sum
	lastCol[0].offset(column=1).value = enthalpy_sum
	lastCol[0].offset(column=1).fill = YELLOW_HIGHLIGHT

	entropy_flow = find_row_with_key(worksheet, "Entropy Flow")
	entropy_sum = 0 #Initialize counter	
	for col in worksheet.iter_cols(min_row=entropy_flow, max_row=entropy_flow, min_col=3, max_col=worksheet.max_column):
		for cell in col:
			in_out_row = str(cell.column_letter) + "1"
			if worksheet[in_out_row].value == "In": #Check if the first row in this col has In or Out
				entropy_sum += cell.value 			#If In, Add to the sum
			if worksheet[in_out_row].value == "Out":
				entropy_sum -= cell.value 			#If out, Subtract from sum
	lastCol[0].offset(row=1,column=1).value = entropy_sum
	lastCol[0].offset(row=1,column=1).fill = YELLOW_HIGHLIGHT

	exergy_flow = find_row_with_key(worksheet, "Exergy Flow")
	exergy_sum = 0 #Initialize counter	
	for col in worksheet.iter_cols(min_row=exergy_flow, max_row=exergy_flow, min_col=3, max_col=worksheet.max_column):
		for cell in col:
			in_out_row = str(cell.column_letter) + "1"
			if worksheet[in_out_row].value == "In": #Check if the first row in this col has In or Out
				exergy_sum += cell.value 			#If In, Add to the sum
			if worksheet[in_out_row].value == "Out":
				exergy_sum -= cell.value 			#If out, Subtract from sum
	lastCol[0].offset(row=2,column=1).value = exergy_sum
	lastCol[0].offset(row=2,column=1).fill = YELLOW_HIGHLIGHT
	worksheet[(str((lastCol[0].offset(column=1)).column_letter) + "1")].value = "Balances" #Convoluted, just add title to cell

def step_six(worksheet):
	#Lesson, do not use array for deleting rows because they change dynamically per each deletion
	worksheet.delete_rows(1, 2)
	worksheet.delete_rows(find_row_with_key(worksheet, "Maximum Relative Error"), 1)
	worksheet.delete_rows(find_row_with_key(worksheet, "Cost Flow"), 1)
	worksheet.delete_rows(find_row_with_key(worksheet, "MIXED Substream"), 1)
	worksheet.delete_rows(find_row_with_key(worksheet, "Mass Vapor Fraction"), 1)
	worksheet.delete_rows(find_row_with_key(worksheet, "Mass Liquid Fraction"), 1)
	worksheet.delete_rows(find_row_with_key(worksheet, "Mass Solid Fraction"), 1)
	worksheet.delete_rows(find_row_with_key(worksheet, "Mass Enthalpy"), 1)
	worksheet.delete_rows(find_row_with_key(worksheet, "Mass Entropy"), 1)
	worksheet.delete_rows(find_row_with_key(worksheet, "Mass Density"), 1)
	worksheet.delete_rows(find_row_with_key(worksheet, "Molar Liquid Fraction"), 1)
	worksheet.delete_rows(find_row_with_key(worksheet, "Molar Solid Fraction"), 1)
	removeRowsBelow(worksheet)
	removeZeroRows(worksheet)
	enthalpy_flow = find_row_with_key(worksheet, "Enthalpy Flow")
	enthalpy_flow_units = "B" + str(enthalpy_flow)
	if(worksheet[enthalpy_flow_units].value == "W"):
		print("Enthalpy Flow in Watts, converting to MW")
		for col in worksheet.iter_cols(min_col=3, max_col=worksheet.max_column, 
				min_row=enthalpy_flow, max_row=enthalpy_flow):
			if(col[0].value != None):
				newVal = col[0].value / 1000000
				col[0].value = newVal

def add_text(worksheet, title, overall_text):
	worksheet['A1'] = title
	cnt = 0
	name_array= []
	for col in worksheet.iter_rows(min_row=3, max_row=3):
		for cell in col:
			if(cell.value == "Name"):
				name_array.append(cell.col_idx)
	
	for x in name_array:
		cnt = 0
		for col in worksheet.iter_rows(min_row=64, max_row=110, min_col=x, max_col=x):
			for cell in col:
				if (overall_text[cnt] == "Inlet Stream 1 Name"
					or overall_text[cnt] == "Inlet Stream 2 Name"
					or overall_text[cnt] == "Inlet Stream 3 Name"
					or overall_text[cnt] == "Inlet Stream 4 Name"):
					cell.fill = YELLOW_HIGHLIGHT
				if (overall_text[cnt] == "Outlet Stream 1 Name"
					or overall_text[cnt] == "Outlet Stream 2 Name"
					or overall_text[cnt] == "Outlet Stream 3 Name"
					or overall_text[cnt] == "Outlet Stream 4 Name"
					or overall_text[cnt] == "Outlet Stream 5 Name"
					or overall_text[cnt] == "Outlet Stream 6 Name"):
					cell.fill = ORANGE_HIGHLIGHT
				if (overall_text[cnt] == "Mass balance kg/hr"
					or overall_text[cnt] == "Energy Balance MW"
					or overall_text[cnt] == "Entropy Generation kW/K"):
					cell.fill=GREEN_HIGHLIGHT
				cell.value = overall_text[cnt]
			cnt += 1

	for col in worksheet.iter_rows(min_row=3, max_row=3):
		for cell in col:
			if(cell.value != "Name" and cell.value != None):
				temp = cell.value
				thisCell = cell.offset(row=62)
				thisCell.value = temp

	block_name_array = []
	idx = 0
	for col in worksheet.iter_rows(min_row=2, max_row=2):
		for cell in col:
			if(cell.value != None):
				block_name_array.append(cell.value)
	for col in worksheet.iter_cols(min_row=64, max_row=64):
		for cell in col:
			if cell.value == "Block Type":
				cell.offset(column=1).value = block_name_array[idx]
				idx += 1


def copy_worksheet(workbook, sheetName):
	source = workbook.active
	target = workbook.copy_worksheet(source)
	target.title = sheetName
	return target

def step_seven(worksheet,title):
	addTitle(worksheet,title)
	addInOutRows(worksheet)
	entropyCalculations(worksheet)
	freezeCells = worksheet['C7']
	worksheet.freeze_panes = freezeCells

def step_eight(worksheet):
	freezeCells = worksheet['C7']
	worksheet.freeze_panes = freezeCells
	removeColumns(worksheet)
	addInOutValues(worksheet)

def step_nine(worksheet):
	calculate_balance(worksheet)

#Looks at the to row, returns array of tuples with format as follows:
#	(to-row-name, stream-name)
def prepare_for_overall_inlet(worksheet):
	return_arr = []
	to_row = find_row_with_key(worksheet,"To")
	for col in worksheet.iter_cols(min_row=to_row, max_row=to_row,min_col=3):
		for cell in col:
			if cell.value != None:
				return_arr.append((cell.value, (cell.offset(row=-2).value)))
	print(return_arr)
	return return_arr

#Looks at the from row, returns array of tuples with format as follows:
#	(from-row-name, stream-name)
def prepare_for_overall_outlet(worksheet):
	return_arr = []
	from_row = find_row_with_key(worksheet,"From")
	for col in worksheet.iter_cols(min_row=from_row, max_row=from_row,min_col=3):
		for cell in col:
			if cell.value != None:
				return_arr.append((cell.value, (cell.offset(row=-2).value)))
	return return_arr

def step_twelve(worksheet, inlet_array):
	for col in worksheet.iter_cols(min_row=65, max_row=65,min_col=2):
		for cell in col:
			if cell.value != "Block Name" and cell.value != None:
				for x in inlet_array:
					curr = x[0]
					if cell.value == curr:
						print("here")
						thisCell = cell.offset(row=1)
						thisCell.value = x[1]

def main():
	with tqdm(total=100, file=sys.stdout) as pbar:
		for i in range(1):
			#print("Main Source file called")
			inputData = get_config_variables()
			streamWorkbook = inputData["streamBookName"]
			#print("Working on: " + str(streamWorkbook))
			wb_stream = openpyxl.load_workbook(streamWorkbook)
			modifiedWS = copy_worksheet(wb_stream, "Aspen Data Tables Modified")
			#Begin work on streams workbook
			step_six(modifiedWS)
			pbar.update(25)
			step_seven(modifiedWS, inputData["streamTitle"]) 
			wb_stream.save(streamWorkbook)
			overall = wb_stream.copy_worksheet(modifiedWS)
			overall.title = "Overall"
			pbar.update(25)
			step_eight(overall)
			pbar.update(25)
			step_nine(overall)
			wb_stream.save(streamWorkbook)
			pbar.update(25)

	with tqdm(total=100, file=sys.stdout) as pbar:
		for i in range(1):
			#Begin work on models workbook 
			inputData = get_config_variables()
			modelWorkbook = inputData["modelBookName"]
			print("Working on:" + str(modelWorkbook))
			wb_block = openpyxl.load_workbook(modelWorkbook)
			overallTitle = inputData["modelTitle"]
			overallWS = copy_worksheet(wb_block, "Overall")
			overall_text_add = inputData["overall_text_add"]
			add_text(overallWS, overallTitle, overall_text_add)
			wb_block.save(modelWorkbook)
			pbar.update(50)
			inlet_array = prepare_for_overall_inlet(overall)
			step_twelve(overallWS, inlet_array)
			wb_block.save(modelWorkbook)
			pbar.update(50)

if __name__ == '__main__':
	main()

#Big question: Follow logic as such - Radfrac, in example given, under C-301 has S35D. 
# Find this in streams and it is one of the deleted columns because both to and from are filled
# Where is the error in this? (12)

# Big Question: Did I calculate the correct rows (accoridng to example yes but the instructions are confusing)
# What are the check that should be performed here?