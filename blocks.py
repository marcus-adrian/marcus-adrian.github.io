#TODO: Refactor function names and variables (low-hyphen instead of camelcase)

import yaml
import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
import time
import sys
from tqdm import tqdm

YELLOW_HIGHLIGHT = PatternFill(start_color='ffff00',
				   end_color='ffff00',
				   fill_type='solid')
GREEN_HIGHLIGHT = PatternFill(start_color='90ee90',
				   end_color='90ee90',
				   fill_type='solid')
ORANGE_HIGHLIGHT = PatternFill(start_color='ffa500',
				   end_color='ffa500',
				   fill_type='solid')

#Retrieves data from the YAML file and returns a dictionary with the vals.
def get_config_variables():
	with open('config.yaml') as f:
		data = yaml.load(f, Loader=yaml.FullLoader)
		return data

def add_text(worksheet, title, overall_text):
	worksheet['A1'] = title
	cnt = 0
	name_array= []
	for col in worksheet.iter_rows(min_row=3, max_row=3):
		for cell in col:
			if(cell.value == "Name"):
				name_array.append(cell.col_idx)
	
	for x in name_array:
		cnt = 0
		for col in worksheet.iter_rows(min_row=64, max_row=110, min_col=x, max_col=x):
			for cell in col:
				if (overall_text[cnt] == "Inlet Stream 1 Name"
					or overall_text[cnt] == "Inlet Stream 2 Name"
					or overall_text[cnt] == "Inlet Stream 3 Name"
					or overall_text[cnt] == "Inlet Stream 4 Name"):
					cell.fill = YELLOW_HIGHLIGHT
				if (overall_text[cnt] == "Outlet Stream 1 Name"
					or overall_text[cnt] == "Outlet Stream 2 Name"
					or overall_text[cnt] == "Outlet Stream 3 Name"
					or overall_text[cnt] == "Outlet Stream 4 Name"
					or overall_text[cnt] == "Outlet Stream 5 Name"
					or overall_text[cnt] == "Outlet Stream 6 Name"):
					cell.fill = ORANGE_HIGHLIGHT
				if (overall_text[cnt] == "Mass balance kg/hr"
					or overall_text[cnt] == "Energy Balance MW"
					or overall_text[cnt] == "Entropy Generation kW/K"):
					cell.fill=GREEN_HIGHLIGHT
				cell.value = overall_text[cnt]
			cnt += 1

	for col in worksheet.iter_rows(min_row=3, max_row=3):
		for cell in col:
			if(cell.value != "Name" and cell.value != None):
				temp = cell.value
				thisCell = cell.offset(row=62)
				thisCell.value = temp

	block_name_array = []
	idx = 0
	for col in worksheet.iter_rows(min_row=2, max_row=2):
		for cell in col:
			if(cell.value != None):
				block_name_array.append(cell.value)
	for col in worksheet.iter_cols(min_row=64, max_row=64):
		for cell in col:
			if cell.value == "Block Type":
				cell.offset(column=1).value = block_name_array[idx]
				idx += 1


def copy_worksheet(workbook, sheetName):
	source = workbook.active
	target = workbook.copy_worksheet(source)
	target.title = sheetName
	return target

def main():
	with tqdm(total=100, file=sys.stdout) as pbar:
		for i in range(1):
				#Begin work on models workbook 
				inputData = get_config_variables()
				modelWorkbook = inputData["modelBookName"]
				print("Working on:" + str(modelWorkbook))
				wb = openpyxl.load_workbook(modelWorkbook)
				overallTitle = inputData["modelTitle"]
				overallWS = copy_worksheet(wb, "Overall")
				overall_text_add = inputData["overall_text_add"]
				add_text(overallWS, overallTitle, overall_text_add)
				wb.save(modelWorkbook)
				pbar.update(100)

				"""
				writeRow = writeSections(overallWS) #Pretty bad coding here, could use global but quick fix
				writeOtherSections(overallWS,writeRow)
				wb.save(modelWorkbook)
				"""

if __name__ == '__main__':
	main()